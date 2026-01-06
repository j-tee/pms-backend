"""
Farmer Analytics Export Views

Generates Excel, PDF, and CSV exports of farmer analytics data.
"""

import io
import csv
from datetime import datetime
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, 
    Image, PageBreak, HRFlowable
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

from farms.models import Farm
from .services.farmer_analytics import FarmerAnalyticsService


class BaseExportView(APIView):
    """Base class for export views"""
    permission_classes = [IsAuthenticated]
    
    def get_farm(self, request):
        """Get farm for authenticated user"""
        try:
            return Farm.objects.get(owner=request.user)
        except Farm.DoesNotExist:
            return None
    
    def get_days(self, request):
        """Get days parameter from request"""
        try:
            days = int(request.query_params.get('days', 30))
            return min(max(days, 1), 365)
        except (ValueError, TypeError):
            return 30


class ExportAnalyticsExcelView(BaseExportView):
    """Export farmer analytics to Excel"""
    
    def get(self, request):
        farm = self.get_farm(request)
        if not farm:
            return Response(
                {'error': 'No farm found for this user', 'code': 'NO_FARM'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        days = self.get_days(request)
        service = FarmerAnalyticsService(farm)
        
        # Get all analytics data
        analytics = service.get_full_analytics(days)
        
        # Create workbook
        wb = Workbook()
        
        # Style definitions
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # === SUMMARY SHEET ===
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._create_summary_sheet(ws_summary, analytics, farm, days, header_font, header_fill, title_font, thin_border)
        
        # === PRODUCTION SHEET ===
        ws_production = wb.create_sheet("Production")
        self._create_production_sheet(ws_production, analytics.get('production', {}), header_font, header_fill, thin_border)
        
        # === FLOCK HEALTH SHEET ===
        ws_health = wb.create_sheet("Flock Health")
        self._create_health_sheet(ws_health, analytics.get('flock_health', {}), header_font, header_fill, thin_border)
        
        # === FINANCIAL SHEET ===
        ws_financial = wb.create_sheet("Financial")
        self._create_financial_sheet(ws_financial, analytics.get('financial', {}), header_font, header_fill, thin_border)
        
        # === FEED SHEET ===
        ws_feed = wb.create_sheet("Feed")
        self._create_feed_sheet(ws_feed, analytics.get('feed', {}), header_font, header_fill, thin_border)
        
        # === MARKETPLACE SHEET ===
        if analytics.get('marketplace', {}).get('enabled'):
            ws_marketplace = wb.create_sheet("Marketplace")
            self._create_marketplace_sheet(ws_marketplace, analytics.get('marketplace', {}), header_font, header_fill, thin_border)
        
        # Create response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"farm_analytics_{farm.farm_id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    def _create_summary_sheet(self, ws, analytics, farm, days, header_font, header_fill, title_font, border):
        """Create summary sheet"""
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = f"Farm Analytics Report - {farm.farm_name}"
        ws[f'A{row}'].font = Font(bold=True, size=16)
        row += 1
        
        ws[f'A{row}'] = f"Period: Last {days} days | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        row += 2
        
        # Farm Info
        ws[f'A{row}'] = "Farm Information"
        ws[f'A{row}'].font = title_font
        row += 1
        
        farm_info = analytics.get('farm', {})
        info_data = [
            ('Farm Name', farm_info.get('farm_name', '')),
            ('Farm ID', farm_info.get('farm_number', '')),
            ('Constituency', farm_info.get('constituency', '')),
            ('Production Type', farm_info.get('primary_production_type', '')),
            ('Bird Capacity', farm_info.get('total_bird_capacity', 0)),
            ('Current Birds', farm_info.get('current_bird_count', 0)),
            ('Capacity Utilization', f"{farm_info.get('capacity_utilization', 0)}%"),
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Key Metrics
        ws[f'A{row}'] = "Key Metrics"
        ws[f'A{row}'].font = title_font
        row += 1
        
        production = analytics.get('production', {})
        financial = analytics.get('financial', {})
        health = analytics.get('flock_health', {})
        
        metrics = [
            ('Total Eggs Collected', production.get('summary', {}).get('total_eggs', 0)),
            ('Average Daily Production', production.get('summary', {}).get('avg_daily_production', 0)),
            ('Production Rate', f"{production.get('summary', {}).get('production_rate_percent', 0)}%"),
            ('Mortality Rate', f"{health.get('summary', {}).get('mortality_rate_period', 0)}%"),
            ('Total Revenue', f"GHS {financial.get('summary', {}).get('total_revenue', 0):,.2f}"),
            ('Total Expenses', f"GHS {financial.get('summary', {}).get('total_expenses', 0):,.2f}"),
            ('Gross Profit', f"GHS {financial.get('summary', {}).get('gross_profit', 0):,.2f}"),
            ('Profit Margin', f"{financial.get('summary', {}).get('profit_margin_percent', 0)}%"),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
    
    def _create_production_sheet(self, ws, production, header_font, header_fill, border):
        """Create production sheet"""
        row = 1
        
        # Title
        ws[f'A{row}'] = "Production Analytics"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # Summary
        summary = production.get('summary', {})
        ws[f'A{row}'] = "Summary"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        summary_data = [
            ('Total Eggs', summary.get('total_eggs', 0)),
            ('Average Daily', summary.get('avg_daily_production', 0)),
            ('Production Rate', f"{summary.get('production_rate_percent', 0)}%"),
            ('Laying Birds', summary.get('total_laying_birds', 0)),
            ('Eggs per Bird', summary.get('eggs_per_bird', 0)),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Quality breakdown
        ws[f'A{row}'] = "Egg Quality Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        quality = production.get('quality', {})
        quality_data = [
            ('Good Eggs', quality.get('good', 0)),
            ('Broken Eggs', quality.get('broken', 0)),
            ('Dirty Eggs', quality.get('dirty', 0)),
            ('Small Eggs', quality.get('small', 0)),
            ('Soft Shell', quality.get('soft_shell', 0)),
            ('Good Percentage', f"{quality.get('good_percentage', 0)}%"),
        ]
        
        for label, value in quality_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Daily trend
        daily_trend = production.get('daily_trend', [])
        if daily_trend:
            ws[f'A{row}'] = "Daily Production Trend"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            # Headers
            headers = ['Date', 'Eggs Collected', 'Good', 'Broken']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            
            for day in daily_trend:
                ws.cell(row=row, column=1, value=str(day.get('production_date', '')))
                ws.cell(row=row, column=2, value=day.get('eggs', 0))
                ws.cell(row=row, column=3, value=day.get('good', 0))
                ws.cell(row=row, column=4, value=day.get('broken', 0))
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_health_sheet(self, ws, health, header_font, header_fill, border):
        """Create flock health sheet"""
        row = 1
        
        ws[f'A{row}'] = "Flock Health & Mortality"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # Summary
        summary = health.get('summary', {})
        summary_data = [
            ('Current Birds', summary.get('current_bird_count', 0)),
            ('Initial Birds', summary.get('initial_bird_count', 0)),
            ('Period Deaths', summary.get('period_deaths', 0)),
            ('Mortality Rate', f"{summary.get('mortality_rate_period', 0)}%"),
            ('Survival Rate', f"{summary.get('survival_rate', 0)}%"),
            ('Avg Daily Mortality', summary.get('avg_daily_mortality', 0)),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Causes breakdown
        ws[f'A{row}'] = "Mortality by Cause"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        causes = health.get('causes_breakdown', {})
        for cause, count in causes.items():
            ws[f'A{row}'] = cause.replace('_', ' ').title()
            ws[f'B{row}'] = count
            row += 1
        
        row += 1
        
        # Flock details
        flocks = health.get('flocks', [])
        if flocks:
            ws[f'A{row}'] = "Flock Details"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            headers = ['Flock #', 'Type', 'Breed', 'Current', 'Initial', 'Deaths', 'Mortality %']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            
            for flock in flocks:
                ws.cell(row=row, column=1, value=flock.get('flock_number', ''))
                ws.cell(row=row, column=2, value=flock.get('flock_type', ''))
                ws.cell(row=row, column=3, value=flock.get('breed', ''))
                ws.cell(row=row, column=4, value=flock.get('current_count', 0))
                ws.cell(row=row, column=5, value=flock.get('initial_count', 0))
                ws.cell(row=row, column=6, value=flock.get('mortality_count', 0))
                ws.cell(row=row, column=7, value=f"{flock.get('mortality_rate', 0)}%")
                row += 1
        
        # Adjust columns
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_financial_sheet(self, ws, financial, header_font, header_fill, border):
        """Create financial sheet"""
        row = 1
        
        ws[f'A{row}'] = "Financial Analytics"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # Summary
        summary = financial.get('summary', {})
        ws[f'A{row}'] = "Summary"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        summary_data = [
            ('Total Revenue', f"GHS {summary.get('total_revenue', 0):,.2f}"),
            ('Total Expenses', f"GHS {summary.get('total_expenses', 0):,.2f}"),
            ('Gross Profit', f"GHS {summary.get('gross_profit', 0):,.2f}"),
            ('Profit Margin', f"{summary.get('profit_margin_percent', 0)}%"),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Revenue breakdown
        ws[f'A{row}'] = "Revenue Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        revenue = financial.get('revenue_breakdown', {})
        eggs = revenue.get('eggs', {})
        birds = revenue.get('birds', {})
        marketplace = revenue.get('marketplace', {})
        
        revenue_data = [
            ('Egg Sales (Gross)', f"GHS {eggs.get('gross', 0):,.2f}"),
            ('Egg Sales (Net)', f"GHS {eggs.get('net', 0):,.2f}"),
            ('Egg Transactions', eggs.get('transactions', 0)),
            ('Bird Sales (Gross)', f"GHS {birds.get('gross', 0):,.2f}"),
            ('Bird Sales (Net)', f"GHS {birds.get('net', 0):,.2f}"),
            ('Birds Sold', birds.get('birds_sold', 0)),
            ('Marketplace Revenue', f"GHS {marketplace.get('gross', 0):,.2f}"),
            ('Marketplace Orders', marketplace.get('orders', 0)),
        ]
        
        for label, value in revenue_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Expenses breakdown
        ws[f'A{row}'] = "Expenses Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        expenses = financial.get('expenses_breakdown', {})
        for expense_type, amount in expenses.items():
            ws[f'A{row}'] = expense_type.title()
            ws[f'B{row}'] = f"GHS {amount:,.2f}"
            row += 1
        
        # Adjust columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_feed_sheet(self, ws, feed, header_font, header_fill, border):
        """Create feed analytics sheet"""
        row = 1
        
        ws[f'A{row}'] = "Feed Analytics"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # Summary
        summary = feed.get('summary', {})
        summary_data = [
            ('Total Feed Consumed (kg)', summary.get('total_feed_consumed_kg', 0)),
            ('Total Feed Cost', f"GHS {summary.get('total_feed_cost', 0):,.2f}"),
            ('Avg Daily Consumption (kg)', summary.get('avg_daily_consumption_kg', 0)),
            ('Feed per Bird (grams/day)', summary.get('feed_per_bird_grams', 0)),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Efficiency
        ws[f'A{row}'] = "Feed Efficiency"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        efficiency = feed.get('efficiency', {})
        efficiency_data = [
            ('FCR (kg per dozen eggs)', efficiency.get('fcr_kg_per_dozen_eggs', 0)),
            ('Cost per Egg', f"GHS {efficiency.get('cost_per_egg', 0):.4f}"),
            ('Cost per Crate', f"GHS {efficiency.get('cost_per_crate', 0):.2f}"),
        ]
        
        for label, value in efficiency_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Inventory
        ws[f'A{row}'] = "Feed Inventory"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        inventory = feed.get('inventory', {})
        inventory_data = [
            ('Current Stock (kg)', inventory.get('current_stock_kg', 0)),
            ('Stock Value', f"GHS {inventory.get('stock_value', 0):,.2f}"),
            ('Days Remaining', inventory.get('days_remaining', 0)),
            ('Reorder Alert', 'Yes' if inventory.get('reorder_alert') else 'No'),
        ]
        
        for label, value in inventory_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Adjust columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_marketplace_sheet(self, ws, marketplace, header_font, header_fill, border):
        """Create marketplace analytics sheet"""
        row = 1
        
        ws[f'A{row}'] = "Marketplace Analytics"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # Summary
        summary = marketplace.get('summary', {})
        summary_data = [
            ('Total Orders', summary.get('total_orders', 0)),
            ('Completed Orders', summary.get('completed_orders', 0)),
            ('Pending Orders', summary.get('pending_orders', 0)),
            ('Cancelled Orders', summary.get('cancelled_orders', 0)),
            ('Total Revenue', f"GHS {summary.get('total_revenue', 0):,.2f}"),
            ('Avg Order Value', f"GHS {summary.get('avg_order_value', 0):,.2f}"),
            ('Completion Rate', f"{summary.get('completion_rate', 0)}%"),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Customers
        ws[f'A{row}'] = "Customer Analytics"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        customers = marketplace.get('customers', {})
        customer_data = [
            ('Unique Customers', customers.get('unique_customers', 0)),
            ('Repeat Customers', customers.get('repeat_customers', 0)),
            ('Repeat Rate', f"{customers.get('repeat_rate', 0)}%"),
        ]
        
        for label, value in customer_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Top sellers
        products = marketplace.get('products', {})
        top_sellers = products.get('top_sellers', [])
        if top_sellers:
            ws[f'A{row}'] = "Top Selling Products"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            headers = ['Product', 'Category', 'Quantity Sold', 'Revenue', 'Orders']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            row += 1
            
            for product in top_sellers:
                ws.cell(row=row, column=1, value=product.get('name', ''))
                ws.cell(row=row, column=2, value=product.get('category', ''))
                ws.cell(row=row, column=3, value=product.get('quantity_sold', 0))
                ws.cell(row=row, column=4, value=f"GHS {product.get('revenue', 0):,.2f}")
                ws.cell(row=row, column=5, value=product.get('orders', 0))
                row += 1
        
        # Adjust columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 10


class ExportAnalyticsPDFView(BaseExportView):
    """Export farmer analytics to PDF"""
    
    def get(self, request):
        farm = self.get_farm(request)
        if not farm:
            return Response(
                {'error': 'No farm found for this user', 'code': 'NO_FARM'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        days = self.get_days(request)
        service = FarmerAnalyticsService(farm)
        
        # Get all analytics data
        analytics = service.get_full_analytics(days)
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2E7D32'),
            spaceAfter=10,
            spaceBefore=15
        )
        normal_style = styles['Normal']
        
        # Build document content
        elements = []
        
        # Title
        elements.append(Paragraph(f"Farm Analytics Report", title_style))
        elements.append(Paragraph(f"<b>{farm.farm_name}</b>", ParagraphStyle('FarmName', parent=styles['Normal'], fontSize=14, alignment=TA_CENTER)))
        elements.append(Paragraph(f"Period: Last {days} days | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ParagraphStyle('Date', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER)))
        elements.append(Spacer(1, 20))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#2E7D32')))
        elements.append(Spacer(1, 15))
        
        # Farm Information
        elements.append(Paragraph("Farm Information", heading_style))
        farm_info = analytics.get('farm', {})
        farm_data = [
            ['Farm Name', farm_info.get('farm_name', '')],
            ['Farm ID', farm_info.get('farm_number', '')],
            ['Constituency', farm_info.get('constituency', '')],
            ['Production Type', farm_info.get('primary_production_type', '')],
            ['Bird Capacity', str(farm_info.get('total_bird_capacity', 0))],
            ['Current Birds', str(farm_info.get('current_bird_count', 0))],
            ['Capacity Utilization', f"{farm_info.get('capacity_utilization', 0)}%"],
        ]
        
        farm_table = Table(farm_data, colWidths=[4*cm, 10*cm])
        farm_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(farm_table)
        elements.append(Spacer(1, 15))
        
        # Key Metrics Summary
        elements.append(Paragraph("Key Performance Metrics", heading_style))
        
        production = analytics.get('production', {})
        financial = analytics.get('financial', {})
        health = analytics.get('flock_health', {})
        
        metrics_data = [
            ['Metric', 'Value'],
            ['Total Eggs Collected', str(production.get('summary', {}).get('total_eggs', 0))],
            ['Average Daily Production', str(production.get('summary', {}).get('avg_daily_production', 0))],
            ['Production Rate', f"{production.get('summary', {}).get('production_rate_percent', 0)}%"],
            ['Mortality Rate', f"{health.get('summary', {}).get('mortality_rate_period', 0)}%"],
            ['Total Revenue', f"GHS {financial.get('summary', {}).get('total_revenue', 0):,.2f}"],
            ['Total Expenses', f"GHS {financial.get('summary', {}).get('total_expenses', 0):,.2f}"],
            ['Gross Profit', f"GHS {financial.get('summary', {}).get('gross_profit', 0):,.2f}"],
            ['Profit Margin', f"{financial.get('summary', {}).get('profit_margin_percent', 0)}%"],
        ]
        
        metrics_table = Table(metrics_data, colWidths=[7*cm, 7*cm])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E7D32')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(metrics_table)
        
        # Production Details
        elements.append(PageBreak())
        elements.append(Paragraph("Production Analytics", heading_style))
        
        quality = production.get('quality', {})
        quality_data = [
            ['Egg Quality', 'Count'],
            ['Good Eggs', str(quality.get('good', 0))],
            ['Broken Eggs', str(quality.get('broken', 0))],
            ['Dirty Eggs', str(quality.get('dirty', 0))],
            ['Small Eggs', str(quality.get('small', 0))],
            ['Soft Shell', str(quality.get('soft_shell', 0))],
            ['Good %', f"{quality.get('good_percentage', 0)}%"],
        ]
        
        quality_table = Table(quality_data, colWidths=[7*cm, 7*cm])
        quality_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4CAF50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(quality_table)
        elements.append(Spacer(1, 15))
        
        # Flock Health
        elements.append(Paragraph("Flock Health & Mortality", heading_style))
        
        health_summary = health.get('summary', {})
        causes = health.get('causes_breakdown', {})
        
        health_data = [
            ['Health Metric', 'Value'],
            ['Current Bird Count', str(health_summary.get('current_bird_count', 0))],
            ['Period Deaths', str(health_summary.get('period_deaths', 0))],
            ['Mortality Rate', f"{health_summary.get('mortality_rate_period', 0)}%"],
            ['Survival Rate', f"{health_summary.get('survival_rate', 0)}%"],
        ]
        
        health_table = Table(health_data, colWidths=[7*cm, 7*cm])
        health_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FF9800')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(health_table)
        elements.append(Spacer(1, 10))
        
        # Mortality causes
        if causes:
            causes_data = [['Cause', 'Deaths']]
            for cause, count in causes.items():
                if count > 0:
                    causes_data.append([cause.replace('_', ' ').title(), str(count)])
            
            if len(causes_data) > 1:
                causes_table = Table(causes_data, colWidths=[7*cm, 7*cm])
                causes_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#FFC107')),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                elements.append(causes_table)
        
        # Financial Details
        elements.append(PageBreak())
        elements.append(Paragraph("Financial Analytics", heading_style))
        
        revenue = financial.get('revenue_breakdown', {})
        expenses = financial.get('expenses_breakdown', {})
        
        financial_data = [
            ['Category', 'Amount (GHS)'],
            ['Egg Sales', f"{revenue.get('eggs', {}).get('gross', 0):,.2f}"],
            ['Bird Sales', f"{revenue.get('birds', {}).get('gross', 0):,.2f}"],
            ['Marketplace', f"{revenue.get('marketplace', {}).get('gross', 0):,.2f}"],
            ['', ''],
            ['Feed Expenses', f"{expenses.get('feed', 0):,.2f}"],
            ['Medication', f"{expenses.get('medication', 0):,.2f}"],
            ['Vaccination', f"{expenses.get('vaccination', 0):,.2f}"],
        ]
        
        financial_table = Table(financial_data, colWidths=[7*cm, 7*cm])
        financial_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1976D2')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(financial_table)
        elements.append(Spacer(1, 15))
        
        # Feed Analytics
        elements.append(Paragraph("Feed Analytics", heading_style))
        
        feed = analytics.get('feed', {})
        feed_summary = feed.get('summary', {})
        feed_efficiency = feed.get('efficiency', {})
        feed_inventory = feed.get('inventory', {})
        
        feed_data = [
            ['Feed Metric', 'Value'],
            ['Total Consumed (kg)', str(feed_summary.get('total_feed_consumed_kg', 0))],
            ['Total Cost', f"GHS {feed_summary.get('total_feed_cost', 0):,.2f}"],
            ['Avg Daily (kg)', str(feed_summary.get('avg_daily_consumption_kg', 0))],
            ['FCR (kg/dozen)', str(feed_efficiency.get('fcr_kg_per_dozen_eggs', 0))],
            ['Cost per Egg', f"GHS {feed_efficiency.get('cost_per_egg', 0):.4f}"],
            ['Current Stock (kg)', str(feed_inventory.get('current_stock_kg', 0))],
            ['Days Remaining', str(feed_inventory.get('days_remaining', 0))],
        ]
        
        feed_table = Table(feed_data, colWidths=[7*cm, 7*cm])
        feed_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#795548')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(feed_table)
        
        # Marketplace (if enabled)
        marketplace = analytics.get('marketplace', {})
        if marketplace.get('enabled'):
            elements.append(Spacer(1, 15))
            elements.append(Paragraph("Marketplace Analytics", heading_style))
            
            mp_summary = marketplace.get('summary', {})
            mp_data = [
                ['Marketplace Metric', 'Value'],
                ['Total Orders', str(mp_summary.get('total_orders', 0))],
                ['Completed', str(mp_summary.get('completed_orders', 0))],
                ['Pending', str(mp_summary.get('pending_orders', 0))],
                ['Total Revenue', f"GHS {mp_summary.get('total_revenue', 0):,.2f}"],
                ['Avg Order Value', f"GHS {mp_summary.get('avg_order_value', 0):,.2f}"],
                ['Completion Rate', f"{mp_summary.get('completion_rate', 0)}%"],
            ]
            
            mp_table = Table(mp_data, colWidths=[7*cm, 7*cm])
            mp_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9C27B0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            elements.append(mp_table)
        
        # Footer
        elements.append(Spacer(1, 30))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.grey))
        elements.append(Paragraph(
            f"Generated by YEA Poultry Management System | {farm.farm_name} | {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.grey)
        ))
        
        # Build PDF
        doc.build(elements)
        
        # Create response
        buffer.seek(0)
        filename = f"farm_analytics_{farm.farm_id}_{datetime.now().strftime('%Y%m%d')}.pdf"
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class ExportAnalyticsCSVView(BaseExportView):
    """Export farmer analytics to CSV"""
    
    def get(self, request):
        farm = self.get_farm(request)
        if not farm:
            return Response(
                {'error': 'No farm found for this user', 'code': 'NO_FARM'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        days = self.get_days(request)
        section = request.query_params.get('section', 'all')
        service = FarmerAnalyticsService(farm)
        
        # Get analytics data
        analytics = service.get_full_analytics(days)
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        filename = f"farm_analytics_{farm.farm_id}_{section}_{datetime.now().strftime('%Y%m%d')}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        
        # Header info
        writer.writerow(['YEA Poultry Management System - Farm Analytics Export'])
        writer.writerow(['Farm', farm.farm_name])
        writer.writerow(['Farm ID', farm.farm_id])
        writer.writerow(['Period', f'Last {days} days'])
        writer.writerow(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M')])
        writer.writerow([])
        
        if section in ['all', 'summary']:
            self._write_summary_section(writer, analytics)
        
        if section in ['all', 'production']:
            self._write_production_section(writer, analytics.get('production', {}))
        
        if section in ['all', 'health']:
            self._write_health_section(writer, analytics.get('flock_health', {}))
        
        if section in ['all', 'financial']:
            self._write_financial_section(writer, analytics.get('financial', {}))
        
        if section in ['all', 'feed']:
            self._write_feed_section(writer, analytics.get('feed', {}))
        
        if section in ['all', 'marketplace']:
            marketplace = analytics.get('marketplace', {})
            if marketplace.get('enabled'):
                self._write_marketplace_section(writer, marketplace)
        
        if section in ['all', 'daily']:
            self._write_daily_data_section(writer, analytics)
        
        return response
    
    def _write_summary_section(self, writer, analytics):
        """Write summary section"""
        writer.writerow(['=== KEY METRICS SUMMARY ==='])
        writer.writerow(['Metric', 'Value'])
        
        farm = analytics.get('farm', {})
        production = analytics.get('production', {}).get('summary', {})
        financial = analytics.get('financial', {}).get('summary', {})
        health = analytics.get('flock_health', {}).get('summary', {})
        
        metrics = [
            ('Current Bird Count', farm.get('current_bird_count', 0)),
            ('Bird Capacity', farm.get('total_bird_capacity', 0)),
            ('Capacity Utilization %', farm.get('capacity_utilization', 0)),
            ('Total Eggs Collected', production.get('total_eggs', 0)),
            ('Average Daily Eggs', production.get('avg_daily_production', 0)),
            ('Production Rate %', production.get('production_rate_percent', 0)),
            ('Mortality Rate %', health.get('mortality_rate_period', 0)),
            ('Survival Rate %', health.get('survival_rate', 0)),
            ('Total Revenue (GHS)', financial.get('total_revenue', 0)),
            ('Total Expenses (GHS)', financial.get('total_expenses', 0)),
            ('Gross Profit (GHS)', financial.get('gross_profit', 0)),
            ('Profit Margin %', financial.get('profit_margin_percent', 0)),
        ]
        
        for metric, value in metrics:
            writer.writerow([metric, value])
        writer.writerow([])
    
    def _write_production_section(self, writer, production):
        """Write production section"""
        writer.writerow(['=== PRODUCTION ANALYTICS ==='])
        
        # Summary
        writer.writerow(['Production Summary'])
        summary = production.get('summary', {})
        writer.writerow(['Total Eggs', summary.get('total_eggs', 0)])
        writer.writerow(['Average Daily', summary.get('avg_daily_production', 0)])
        writer.writerow(['Production Rate %', summary.get('production_rate_percent', 0)])
        writer.writerow(['Laying Birds', summary.get('total_laying_birds', 0)])
        writer.writerow(['Eggs per Bird', summary.get('eggs_per_bird', 0)])
        writer.writerow([])
        
        # Quality
        writer.writerow(['Egg Quality Breakdown'])
        quality = production.get('quality', {})
        writer.writerow(['Quality', 'Count'])
        writer.writerow(['Good', quality.get('good', 0)])
        writer.writerow(['Broken', quality.get('broken', 0)])
        writer.writerow(['Dirty', quality.get('dirty', 0)])
        writer.writerow(['Small', quality.get('small', 0)])
        writer.writerow(['Soft Shell', quality.get('soft_shell', 0)])
        writer.writerow(['Good Percentage', f"{quality.get('good_percentage', 0)}%"])
        writer.writerow([])
    
    def _write_health_section(self, writer, health):
        """Write flock health section"""
        writer.writerow(['=== FLOCK HEALTH & MORTALITY ==='])
        
        summary = health.get('summary', {})
        writer.writerow(['Health Summary'])
        writer.writerow(['Current Birds', summary.get('current_bird_count', 0)])
        writer.writerow(['Initial Birds', summary.get('initial_bird_count', 0)])
        writer.writerow(['Period Deaths', summary.get('period_deaths', 0)])
        writer.writerow(['Mortality Rate %', summary.get('mortality_rate_period', 0)])
        writer.writerow(['Survival Rate %', summary.get('survival_rate', 0)])
        writer.writerow(['Avg Daily Deaths', summary.get('avg_daily_mortality', 0)])
        writer.writerow([])
        
        # Causes
        writer.writerow(['Mortality by Cause'])
        writer.writerow(['Cause', 'Deaths'])
        causes = health.get('causes_breakdown', {})
        for cause, count in causes.items():
            writer.writerow([cause.replace('_', ' ').title(), count])
        writer.writerow([])
        
        # Flock details
        flocks = health.get('flocks', [])
        if flocks:
            writer.writerow(['Flock Details'])
            writer.writerow(['Flock #', 'Type', 'Breed', 'Current', 'Initial', 'Deaths', 'Mortality %', 'Age (weeks)'])
            for flock in flocks:
                writer.writerow([
                    flock.get('flock_number', ''),
                    flock.get('flock_type', ''),
                    flock.get('breed', ''),
                    flock.get('current_count', 0),
                    flock.get('initial_count', 0),
                    flock.get('mortality_count', 0),
                    flock.get('mortality_rate', 0),
                    flock.get('age_weeks', ''),
                ])
            writer.writerow([])
    
    def _write_financial_section(self, writer, financial):
        """Write financial section"""
        writer.writerow(['=== FINANCIAL ANALYTICS ==='])
        
        summary = financial.get('summary', {})
        writer.writerow(['Financial Summary'])
        writer.writerow(['Total Revenue (GHS)', summary.get('total_revenue', 0)])
        writer.writerow(['Total Expenses (GHS)', summary.get('total_expenses', 0)])
        writer.writerow(['Gross Profit (GHS)', summary.get('gross_profit', 0)])
        writer.writerow(['Profit Margin %', summary.get('profit_margin_percent', 0)])
        writer.writerow([])
        
        # Revenue breakdown
        writer.writerow(['Revenue Breakdown'])
        revenue = financial.get('revenue_breakdown', {})
        eggs = revenue.get('eggs', {})
        birds = revenue.get('birds', {})
        marketplace = revenue.get('marketplace', {})
        
        writer.writerow(['Source', 'Gross (GHS)', 'Net (GHS)', 'Transactions'])
        writer.writerow(['Egg Sales', eggs.get('gross', 0), eggs.get('net', 0), eggs.get('transactions', 0)])
        writer.writerow(['Bird Sales', birds.get('gross', 0), birds.get('net', 0), birds.get('transactions', 0)])
        writer.writerow(['Marketplace', marketplace.get('gross', 0), '', marketplace.get('orders', 0)])
        writer.writerow([])
        
        # Expenses breakdown
        writer.writerow(['Expenses Breakdown'])
        writer.writerow(['Category', 'Amount (GHS)'])
        expenses = financial.get('expenses_breakdown', {})
        for expense_type, amount in expenses.items():
            writer.writerow([expense_type.title(), amount])
        writer.writerow([])
    
    def _write_feed_section(self, writer, feed):
        """Write feed section"""
        writer.writerow(['=== FEED ANALYTICS ==='])
        
        summary = feed.get('summary', {})
        writer.writerow(['Feed Consumption'])
        writer.writerow(['Total Consumed (kg)', summary.get('total_feed_consumed_kg', 0)])
        writer.writerow(['Total Cost (GHS)', summary.get('total_feed_cost', 0)])
        writer.writerow(['Avg Daily (kg)', summary.get('avg_daily_consumption_kg', 0)])
        writer.writerow(['Per Bird (grams/day)', summary.get('feed_per_bird_grams', 0)])
        writer.writerow([])
        
        efficiency = feed.get('efficiency', {})
        writer.writerow(['Feed Efficiency'])
        writer.writerow(['FCR (kg/dozen eggs)', efficiency.get('fcr_kg_per_dozen_eggs', 0)])
        writer.writerow(['Cost per Egg (GHS)', efficiency.get('cost_per_egg', 0)])
        writer.writerow(['Cost per Crate (GHS)', efficiency.get('cost_per_crate', 0)])
        writer.writerow([])
        
        inventory = feed.get('inventory', {})
        writer.writerow(['Feed Inventory'])
        writer.writerow(['Current Stock (kg)', inventory.get('current_stock_kg', 0)])
        writer.writerow(['Stock Value (GHS)', inventory.get('stock_value', 0)])
        writer.writerow(['Days Remaining', inventory.get('days_remaining', 0)])
        writer.writerow(['Reorder Alert', 'Yes' if inventory.get('reorder_alert') else 'No'])
        writer.writerow([])
    
    def _write_marketplace_section(self, writer, marketplace):
        """Write marketplace section"""
        writer.writerow(['=== MARKETPLACE ANALYTICS ==='])
        
        summary = marketplace.get('summary', {})
        writer.writerow(['Marketplace Summary'])
        writer.writerow(['Total Orders', summary.get('total_orders', 0)])
        writer.writerow(['Completed Orders', summary.get('completed_orders', 0)])
        writer.writerow(['Pending Orders', summary.get('pending_orders', 0)])
        writer.writerow(['Cancelled Orders', summary.get('cancelled_orders', 0)])
        writer.writerow(['Total Revenue (GHS)', summary.get('total_revenue', 0)])
        writer.writerow(['Avg Order Value (GHS)', summary.get('avg_order_value', 0)])
        writer.writerow(['Completion Rate %', summary.get('completion_rate', 0)])
        writer.writerow([])
        
        customers = marketplace.get('customers', {})
        writer.writerow(['Customer Analytics'])
        writer.writerow(['Unique Customers', customers.get('unique_customers', 0)])
        writer.writerow(['Repeat Customers', customers.get('repeat_customers', 0)])
        writer.writerow(['Repeat Rate %', customers.get('repeat_rate', 0)])
        writer.writerow([])
        
        # Top sellers
        products = marketplace.get('products', {})
        top_sellers = products.get('top_sellers', [])
        if top_sellers:
            writer.writerow(['Top Selling Products'])
            writer.writerow(['Product', 'Category', 'Quantity Sold', 'Revenue (GHS)', 'Orders'])
            for product in top_sellers:
                writer.writerow([
                    product.get('name', ''),
                    product.get('category', ''),
                    product.get('quantity_sold', 0),
                    product.get('revenue', 0),
                    product.get('orders', 0),
                ])
            writer.writerow([])
    
    def _write_daily_data_section(self, writer, analytics):
        """Write daily production data for detailed analysis"""
        production = analytics.get('production', {})
        daily_trend = production.get('daily_trend', [])
        
        if daily_trend:
            writer.writerow(['=== DAILY PRODUCTION DATA ==='])
            writer.writerow(['Date', 'Eggs Collected', 'Good Eggs', 'Broken Eggs'])
            for day in daily_trend:
                writer.writerow([
                    str(day.get('production_date', '')),
                    day.get('eggs', 0),
                    day.get('good', 0),
                    day.get('broken', 0),
                ])
            writer.writerow([])
