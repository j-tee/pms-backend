"""
National Admin Report Export Views

Export endpoints for generating downloadable PDF and Excel reports
for National Administrators and the Agriculture Minister.
"""

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
import csv
import logging

from .national_admin_views import NationalAdminPermission, BaseNationalAdminView
from .services.national_admin_analytics import NationalAdminAnalyticsService

logger = logging.getLogger(__name__)


# =============================================================================
# EXCEL EXPORT
# =============================================================================

class ExportExecutiveReportExcelView(BaseNationalAdminView):
    """
    GET /api/admin/reports/export/excel/executive/
    
    Export executive dashboard report as Excel workbook.
    Multi-sheet workbook with all key metrics.
    """
    
    def get(self, request):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response(
                {'error': 'Excel export not available. openpyxl not installed.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        
        region, constituency = self.get_scope_params(request)
        service = self.get_service(request, use_cache=True)
        
        # Gather all data
        data = service.get_executive_dashboard(region, constituency)
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        section_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        def style_header_row(ws, row_num, cols):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
        
        # =====================================================================
        # Sheet 1: Executive Summary
        # =====================================================================
        ws = wb.active
        ws.title = "Executive Summary"
        
        # Title
        ws['A1'] = "YEA Poultry Management System - Executive Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')
        
        # Scope
        scope = "National"
        if constituency:
            scope = f"Constituency: {constituency}"
        elif region:
            scope = f"Region: {region}"
        
        ws['A2'] = f"Scope: {scope}"
        ws['A3'] = f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Program Performance Section
        row = 5
        ws.cell(row=row, column=1, value="PROGRAM PERFORMANCE")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        perf = data.get('program_performance', {}).get('summary', {})
        metrics = [
            ("Total Farms", perf.get('total_farms', 0)),
            ("Operational Farms", perf.get('operational_farms', 0)),
            ("Government Supported", perf.get('government_supported', 0)),
            ("Independent Farms", perf.get('independent', 0)),
            ("New This Month", perf.get('new_this_month', 0)),
            ("Growth Rate (%)", perf.get('growth_rate_percent', 0)),
        ]
        
        row += 1
        for label, value in metrics:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Production Section
        row += 1
        ws.cell(row=row, column=1, value="PRODUCTION (Last 30 Days)")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        prod = data.get('production', {}).get('production', {})
        birds = data.get('production', {}).get('birds', {})
        metrics = [
            ("Total Eggs Collected", prod.get('total_eggs', 0)),
            ("Good Eggs", prod.get('good_eggs', 0)),
            ("Egg Quality Rate (%)", prod.get('egg_quality_rate_percent', 0)),
            ("Avg Eggs per Farm", prod.get('avg_eggs_per_farm', 0)),
            ("Total Birds", birds.get('total', 0)),
            ("Capacity Utilization (%)", birds.get('utilization_percent', 0)),
        ]
        
        row += 1
        for label, value in metrics:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Financial Section
        row += 1
        ws.cell(row=row, column=1, value="FINANCIAL METRICS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        fin = data.get('financial', {})
        mkt = fin.get('marketplace', {})
        inc = fin.get('farmer_income', {})
        metrics = [
            ("Total Orders", mkt.get('total_orders', 0)),
            ("Transaction Volume (GHS)", mkt.get('transaction_volume_ghs', 0)),
            ("Active Sellers", mkt.get('active_sellers', 0)),
            ("Gross Farmer Earnings (GHS)", inc.get('gross_earnings_ghs', 0)),
            ("Net Farmer Earnings (GHS)", inc.get('net_earnings_ghs', 0)),
        ]
        
        row += 1
        for label, value in metrics:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Flock Health Section
        row += 1
        ws.cell(row=row, column=1, value="FLOCK HEALTH")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        health = data.get('flock_health', {})
        mort = health.get('mortality', {})
        metrics = [
            ("Total Birds", health.get('total_birds', 0)),
            ("Active Flocks", health.get('active_flocks', 0)),
            ("Total Mortality", mort.get('total', 0)),
            ("Mortality Rate (%)", mort.get('rate_percent', 0)),
        ]
        
        row += 1
        for label, value in metrics:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Farmer Welfare Section
        row += 1
        ws.cell(row=row, column=1, value="FARMER WELFARE & EMPLOYMENT")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        ws.cell(row=row, column=1).fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        welfare = data.get('farmer_welfare', {})
        demo = welfare.get('demographics', {})
        emp = welfare.get('employment_impact', {})
        metrics = [
            ("Total Farmers", demo.get('total_farmers', 0)),
            ("Direct Jobs (Farmers)", emp.get('direct_farmers', 0)),
            ("Estimated Workers", emp.get('estimated_workers', 0)),
            ("Total Jobs Created", emp.get('total_estimated_jobs', 0)),
        ]
        
        row += 1
        for label, value in metrics:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # =====================================================================
        # Sheet 2: Regional Comparison (if national view)
        # =====================================================================
        if not region and not constituency:
            ws2 = wb.create_sheet("Regional Comparison")
            regional_data = service.get_regional_production_comparison()
            
            ws2['A1'] = "Regional Production Comparison"
            ws2['A1'].font = Font(bold=True, size=14)
            
            # Headers
            headers = ['Region', 'Farms', 'Birds', 'Eggs (30d)', 'Mortality (30d)', 'Avg Eggs/Farm']
            for col, header in enumerate(headers, 1):
                cell = ws2.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
            
            # Data
            row = 4
            for region_item in regional_data.get('regions', []):
                ws2.cell(row=row, column=1, value=region_item.get('region', ''))
                ws2.cell(row=row, column=2, value=region_item.get('farms', 0))
                ws2.cell(row=row, column=3, value=region_item.get('birds', 0))
                ws2.cell(row=row, column=4, value=region_item.get('eggs_30d', 0))
                ws2.cell(row=row, column=5, value=region_item.get('mortality_30d', 0))
                ws2.cell(row=row, column=6, value=region_item.get('avg_eggs_per_farm', 0))
                row += 1
            
            # Auto-adjust columns
            for col in range(1, 7):
                ws2.column_dimensions[get_column_letter(col)].width = 18
        
        # =====================================================================
        # Sheet 3: Production Trend
        # =====================================================================
        ws3 = wb.create_sheet("Production Trend")
        
        ws3['A1'] = "Daily Production Trend (Last 30 Days)"
        ws3['A1'].font = Font(bold=True, size=14)
        
        prod_data = data.get('production', {}).get('daily_trend', [])
        
        headers = ['Date', 'Eggs Collected', 'Mortality']
        for col, header in enumerate(headers, 1):
            cell = ws3.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 4
        for item in prod_data:
            ws3.cell(row=row, column=1, value=item.get('date', ''))
            ws3.cell(row=row, column=2, value=item.get('eggs', 0))
            ws3.cell(row=row, column=3, value=item.get('mortality', 0))
            row += 1
        
        for col in range(1, 4):
            ws3.column_dimensions[get_column_letter(col)].width = 18
        
        # =====================================================================
        # Sheet 4: Enrollment Trend
        # =====================================================================
        ws4 = wb.create_sheet("Enrollment Trend")
        
        enrollment = service.get_enrollment_trend(12, region, constituency)
        
        ws4['A1'] = "Monthly Enrollment Trend (Last 12 Months)"
        ws4['A1'].font = Font(bold=True, size=14)
        
        headers = ['Month', 'Total', 'Government', 'Independent']
        for col, header in enumerate(headers, 1):
            cell = ws4.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 4
        for item in enrollment.get('trend', []):
            ws4.cell(row=row, column=1, value=item.get('month', ''))
            ws4.cell(row=row, column=2, value=item.get('total', 0))
            ws4.cell(row=row, column=3, value=item.get('government', 0))
            ws4.cell(row=row, column=4, value=item.get('independent', 0))
            row += 1
        
        for col in range(1, 5):
            ws4.column_dimensions[get_column_letter(col)].width = 15
        
        # Generate response
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"yea_executive_report_{scope.lower().replace(' ', '_').replace(':', '')}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


# =============================================================================
# PDF EXPORT
# =============================================================================

class ExportExecutiveReportPDFView(BaseNationalAdminView):
    """
    GET /api/admin/reports/export/pdf/executive/
    
    Export executive dashboard report as PDF document.
    """
    
    def get(self, request):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, cm
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                PageBreak, Image
            )
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        except ImportError:
            return Response(
                {'error': 'PDF export not available. reportlab not installed.'},
                status=status.HTTP_503_SERVICE_UNAVAILABLE
            )
        
        region, constituency = self.get_scope_params(request)
        service = self.get_service(request, use_cache=True)
        
        # Gather data
        data = service.get_executive_dashboard(region, constituency)
        
        # Scope label
        scope = "National"
        if constituency:
            scope = f"Constituency: {constituency}"
        elif region:
            scope = f"Region: {region}"
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        
        # Styles
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(
            name='MainTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#1F4E79')
        ))
        styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=15,
            spaceAfter=10,
            textColor=colors.HexColor('#2874A6')
        ))
        styles.add(ParagraphStyle(
            name='SubInfo',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.grey
        ))
        
        elements = []
        
        # Title
        elements.append(Paragraph(
            "YEA Poultry Management System",
            styles['MainTitle']
        ))
        elements.append(Paragraph(
            "Executive Report",
            styles['Heading2']
        ))
        elements.append(Paragraph(
            f"Scope: {scope} | Generated: {timezone.now().strftime('%Y-%m-%d %H:%M')}",
            styles['SubInfo']
        ))
        elements.append(Spacer(1, 20))
        
        # Table style
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#EBF5FB')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        # Program Performance Section
        elements.append(Paragraph("Program Performance", styles['SectionTitle']))
        
        perf = data.get('program_performance', {}).get('summary', {})
        perf_data = [
            ['Metric', 'Value'],
            ['Total Farms', str(perf.get('total_farms', 0))],
            ['Operational Farms', str(perf.get('operational_farms', 0))],
            ['Government Supported', str(perf.get('government_supported', 0))],
            ['Independent Farms', str(perf.get('independent', 0))],
            ['New This Month', str(perf.get('new_this_month', 0))],
            ['Growth Rate', f"{perf.get('growth_rate_percent', 0)}%"],
        ]
        
        t = Table(perf_data, colWidths=[4*cm, 3*cm])
        t.setStyle(table_style)
        elements.append(t)
        elements.append(Spacer(1, 15))
        
        # Production Section
        elements.append(Paragraph("Production (Last 30 Days)", styles['SectionTitle']))
        
        prod = data.get('production', {}).get('production', {})
        birds = data.get('production', {}).get('birds', {})
        prod_data = [
            ['Metric', 'Value'],
            ['Total Eggs Collected', f"{prod.get('total_eggs', 0):,}"],
            ['Good Eggs', f"{prod.get('good_eggs', 0):,}"],
            ['Egg Quality Rate', f"{prod.get('egg_quality_rate_percent', 0)}%"],
            ['Avg Eggs per Farm', f"{prod.get('avg_eggs_per_farm', 0):,}"],
            ['Total Birds', f"{birds.get('total', 0):,}"],
            ['Capacity', f"{birds.get('capacity', 0):,}"],
            ['Utilization', f"{birds.get('utilization_percent', 0)}%"],
        ]
        
        t = Table(prod_data, colWidths=[4*cm, 3*cm])
        t.setStyle(table_style)
        elements.append(t)
        elements.append(Spacer(1, 15))
        
        # Financial Section
        elements.append(Paragraph("Financial Metrics", styles['SectionTitle']))
        
        fin = data.get('financial', {})
        mkt = fin.get('marketplace', {})
        inc = fin.get('farmer_income', {})
        fin_data = [
            ['Metric', 'Value'],
            ['Total Orders', str(mkt.get('total_orders', 0))],
            ['Transaction Volume', f"GHS {mkt.get('transaction_volume_ghs', 0):,.2f}"],
            ['Active Sellers', str(mkt.get('active_sellers', 0))],
            ['Gross Farmer Earnings', f"GHS {inc.get('gross_earnings_ghs', 0):,.2f}"],
            ['Net Farmer Earnings', f"GHS {inc.get('net_earnings_ghs', 0):,.2f}"],
        ]
        
        t = Table(fin_data, colWidths=[4*cm, 4*cm])
        t.setStyle(table_style)
        elements.append(t)
        elements.append(Spacer(1, 15))
        
        # Flock Health Section
        elements.append(Paragraph("Flock Health", styles['SectionTitle']))
        
        health = data.get('flock_health', {})
        mort = health.get('mortality', {})
        health_data = [
            ['Metric', 'Value'],
            ['Total Birds', f"{health.get('total_birds', 0):,}"],
            ['Active Flocks', str(health.get('active_flocks', 0))],
            ['Total Mortality', str(mort.get('total', 0))],
            ['Mortality Rate', f"{mort.get('rate_percent', 0)}%"],
        ]
        
        t = Table(health_data, colWidths=[4*cm, 3*cm])
        t.setStyle(table_style)
        elements.append(t)
        elements.append(Spacer(1, 15))
        
        # Employment Impact Section
        elements.append(Paragraph("Employment Impact", styles['SectionTitle']))
        
        welfare = data.get('farmer_welfare', {})
        emp = welfare.get('employment_impact', {})
        emp_data = [
            ['Metric', 'Value'],
            ['Direct Farmers', f"{emp.get('direct_farmers', 0):,}"],
            ['Estimated Workers', f"{emp.get('estimated_workers', 0):,}"],
            ['Total Jobs Created', f"{emp.get('total_estimated_jobs', 0):,}"],
        ]
        
        t = Table(emp_data, colWidths=[4*cm, 3*cm])
        t.setStyle(table_style)
        elements.append(t)
        
        # Regional Comparison (if national)
        if not region and not constituency:
            elements.append(PageBreak())
            elements.append(Paragraph("Regional Production Comparison", styles['SectionTitle']))
            
            regional_data = service.get_regional_production_comparison()
            
            reg_table_data = [['Region', 'Farms', 'Birds', 'Eggs (30d)', 'Mortality']]
            for item in regional_data.get('regions', [])[:10]:  # Top 10
                reg_table_data.append([
                    item.get('region', ''),
                    str(item.get('farms', 0)),
                    f"{item.get('birds', 0):,}",
                    f"{item.get('eggs_30d', 0):,}",
                    str(item.get('mortality_30d', 0)),
                ])
            
            t = Table(reg_table_data, colWidths=[4*cm, 2*cm, 2.5*cm, 2.5*cm, 2*cm])
            t.setStyle(table_style)
            elements.append(t)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        filename = f"yea_executive_report_{scope.lower().replace(' ', '_').replace(':', '')}_{timezone.now().strftime('%Y%m%d')}.pdf"
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response


# =============================================================================
# CSV EXPORT
# =============================================================================

class ExportReportCSVView(BaseNationalAdminView):
    """
    GET /api/admin/reports/export/csv/<report_type>/
    
    Export specific report section as CSV.
    
    Supported report types:
    - production
    - enrollment
    - regional
    - farms
    """
    
    def get(self, request, report_type):
        region, constituency = self.get_scope_params(request)
        service = self.get_service(request, use_cache=True)
        
        response = HttpResponse(content_type='text/csv')
        writer = csv.writer(response)
        
        scope_suffix = ""
        if constituency:
            scope_suffix = f"_{constituency.replace(' ', '_')}"
        elif region:
            scope_suffix = f"_{region.replace(' ', '_')}"
        
        if report_type == 'production':
            filename = f"production_trend{scope_suffix}_{timezone.now().strftime('%Y%m%d')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            data = service.get_production_overview(region, constituency, days=30)
            
            writer.writerow(['Date', 'Eggs Collected', 'Mortality'])
            for item in data.get('daily_trend', []):
                writer.writerow([
                    item.get('date', ''),
                    item.get('eggs', 0),
                    item.get('mortality', 0),
                ])
        
        elif report_type == 'enrollment':
            filename = f"enrollment_trend{scope_suffix}_{timezone.now().strftime('%Y%m%d')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            data = service.get_enrollment_trend(12, region, constituency)
            
            writer.writerow(['Month', 'Total', 'Government', 'Independent'])
            for item in data.get('trend', []):
                writer.writerow([
                    item.get('month', ''),
                    item.get('total', 0),
                    item.get('government', 0),
                    item.get('independent', 0),
                ])
        
        elif report_type == 'regional':
            if region or constituency:
                return Response(
                    {'error': 'Regional comparison only available at national level'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            filename = f"regional_comparison_{timezone.now().strftime('%Y%m%d')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            data = service.get_regional_production_comparison()
            
            writer.writerow(['Region', 'Farms', 'Birds', 'Eggs (30d)', 'Mortality (30d)', 'Avg Eggs/Farm'])
            for item in data.get('regions', []):
                writer.writerow([
                    item.get('region', ''),
                    item.get('farms', 0),
                    item.get('birds', 0),
                    item.get('eggs_30d', 0),
                    item.get('mortality_30d', 0),
                    item.get('avg_eggs_per_farm', 0),
                ])
        
        elif report_type == 'farms':
            filename = f"farms_list{scope_suffix}_{timezone.now().strftime('%Y%m%d')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # Get all farms (paginated)
            page = 1
            page_size = 1000  # Large page for export
            data = service.get_farms_in_scope(region, constituency, page, page_size)
            
            writer.writerow(['Farm ID', 'Farm Name', 'Farmer Name', 'Constituency', 'Status', 'Bird Count', 'Source', 'Created'])
            for farm in data.get('farms', []):
                writer.writerow([
                    farm.get('id', ''),
                    farm.get('farm_name', ''),
                    farm.get('farmer_name', ''),
                    farm.get('constituency', ''),
                    farm.get('status', ''),
                    farm.get('bird_count', 0),
                    farm.get('registration_source', ''),
                    farm.get('created_at', ''),
                ])
        
        else:
            return Response(
                {'error': f'Unknown report type: {report_type}. Valid types: production, enrollment, regional, farms'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return response
